from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from io import BytesIO
from .models import EmpresaOrganizacion, Gremio, OrganizacionCps
from .serializers import EmpresaOrganizacionSerializer, GremioSerializer, OrganizacionCpsSerializer

class GremioViewSet(viewsets.ModelViewSet):
    queryset = Gremio.objects.all()
    serializer_class = GremioSerializer

class OrganizacionCpsViewSet(viewsets.ModelViewSet):
    queryset = OrganizacionCps.objects.all()
    serializer_class = OrganizacionCpsSerializer
    filterset_fields = ['organizacion', 'activa']

class EmpresaOrganizacionViewSet(viewsets.ModelViewSet):
    queryset = EmpresaOrganizacion.objects.all()
    serializer_class = EmpresaOrganizacionSerializer

    @action(detail=True, methods=['get', 'post'])
    def cps(self, request, pk=None):
        """Gestiona los CPS vinculados a una organización específica."""
        org = self.get_object()
        if request.method == 'GET':
            cps = OrganizacionCps.objects.filter(organizacion=org)
            serializer = OrganizacionCpsSerializer(cps, many=True)
            return Response(serializer.data)
        
        if request.method == 'POST':
            data = request.data.copy()
            data['organizacion'] = org.rif
            serializer = OrganizacionCpsSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REGISTRO OPERADORES"
        
        # Headers matching the original Excel exactly
        headers = [
            'MUNICIPIO', 
            'IDENTIFICACIÓN DE LA EMPRESA OPERADORA', 
            'RIF', 
            'REPRESENTANTE LEGAL', 
            'NÚMERO DE TELÉFONO DE CONTACTO',
            'MINIBUS',
            'BUS',
            'OTRO',
            'CERTIFICACIÓN DE PRESTACIÓN DE SERVICIO',
            'URBANAS',
            'SUBURBANAS'
        ]
        ws.append(headers)
        
        for org in EmpresaOrganizacion.objects.all():
            # Get latest CPS if any
            cps_obj = org.cps_registros.filter(activa=True).order_by('-fecha_vencimiento').first()
            cps_text = f"VIGENTE HASTA EL {cps_obj.fecha_vencimiento.strftime('%d/%m/%Y')}" if cps_obj else "NO REGISTRADO"

            ws.append([
                org.municipio.nombre if org.municipio else 'NO ASIGNADO',
                org.razon_social,
                org.rif,
                org.rep_legal_nom,
                org.telefono,
                org.conteo_minibus,
                org.conteo_bus,
                org.conteo_otro,
                cps_text,
                org.rutas_urbanas,
                org.rutas_suburbanas
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=organizaciones.xlsx'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No se subió ningún archivo"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            import re
            from datetime import datetime
            from catalogs.models import TerritorioMunicipio, TipoOrganizacion, TipoCps
            
            wb = openpyxl.load_workbook(file, read_only=True)
            # Find the correct sheet
            sheet = None
            for name in wb.sheetnames:
                if "OPERADORES" in name.upper():
                    sheet = wb[name]
                    break
            
            if not sheet:
                return Response({"error": "No se encontró la pestaña 'REGISTRO OPERADORES'"}, status=status.HTTP_400_BAD_REQUEST)

            # Find header row
            header_row_idx = 1
            for i, row in enumerate(sheet.iter_rows(max_row=20)):
                values = [str(cell.value).strip().upper() if cell.value else "" for cell in row]
                if any("RIF" in v or "IDENTIFICACIÓN" in v or "MUNICIPIO" in v for v in values):
                    header_row_idx = i + 1
                    headers = values
                    break
            
            # Map indices with more flexible matching
            def get_idx(keywords):
                for kw in keywords:
                    for i, h in enumerate(headers):
                        if kw.upper() in h.upper(): return i
                return None

            idx_rif = get_idx(["RIF", "REGISTRO"])
            idx_nombre = get_idx(["IDENTIFICACIÓN", "NOMBRE", "EMPRESA"])
            idx_mun = get_idx(["MUNICIPIO", "SEDE"])
            idx_rep = get_idx(["REPRESENTANTE", "LEGAL"])
            idx_tel = get_idx(["TELÉFONO", "CONTACTO"])
            idx_minibus = get_idx(["MINIBUS"])
            idx_bus = get_idx(["BUS"])
            idx_otro = get_idx(["OTRO"])
            idx_urb = get_idx(["URBANAS"])
            idx_sub = get_idx(["SUBURBANAS"])
            idx_cps = get_idx(["CERTIFICACIÓN", "CPS"])

            if idx_rif is None or idx_nombre is None:
                return Response({"error": "El archivo no tiene el formato esperado (RIF o Nombre faltantes)"}, status=status.HTTP_400_BAD_REQUEST)

            from catalogs.models import TerritorioMunicipio, TipoOrganizacion
            tipo_default, _ = TipoOrganizacion.objects.get_or_create(nombre='LÍNEA/COOPERATIVA')
            
            created_count = 0
            updated_count = 0
            
            for row in sheet.iter_rows(min_row=header_row_idx + 1):
                rif = row[idx_rif].value
                nombre = row[idx_nombre].value
                
                if not rif or not nombre:
                    continue
                
                rif = str(rif).strip()
                nombre = str(nombre).strip()
                telefono = str(row[idx_tel].value).strip() if idx_tel is not None and row[idx_tel].value else ""
                rep_legal = str(row[idx_rep].value).strip() if idx_rep is not None and row[idx_rep].value else ""
                mun_name = str(row[idx_mun].value).strip().upper() if idx_mun is not None and row[idx_mun].value else ""
                
                minibus = int(row[idx_minibus].value) if idx_minibus is not None and isinstance(row[idx_minibus].value, (int, float)) else 0
                bus = int(row[idx_bus].value) if idx_bus is not None and isinstance(row[idx_bus].value, (int, float)) else 0
                otro = int(row[idx_otro].value) if idx_otro is not None and isinstance(row[idx_otro].value, (int, float)) else 0
                urbanas = int(row[idx_urb].value) if idx_urb is not None and isinstance(row[idx_urb].value, (int, float)) else 0
                suburbanas = int(row[idx_sub].value) if idx_sub is not None and isinstance(row[idx_sub].value, (int, float)) else 0

                municipio = None
                if mun_name:
                    municipio = TerritorioMunicipio.objects.filter(nombre__iexact=mun_name).first()

                total_unidades = minibus + bus + otro
                # Infer modality if not present
                modalidad_auto = 'DT10' if total_unidades > 32 else 'DT9'

                org, created = EmpresaOrganizacion.objects.update_or_create(
                    rif=rif,
                    defaults={
                        'razon_social': nombre,
                        'telefono': telefono,
                        'rep_legal_nom': rep_legal,
                        'municipio': municipio,
                        'tipo': tipo_default,
                        'conteo_minibus': minibus,
                        'conteo_bus': bus,
                        'conteo_otro': otro,
                        'cupo_unidades': total_unidades,
                        'modalidad_cps': modalidad_auto,
                        'rutas_urbanas': urbanas,
                        'rutas_suburbanas': suburbanas
                    }
                )
                if created: created_count += 1
                else: updated_count += 1

                # CPS Parsing
                cps_text = str(row[idx_cps].value).strip().upper() if idx_cps is not None and row[idx_cps].value else ""
                if "VIGENTE" in cps_text:
                    # Try to find a date like DD/MM/YY or DD/MM/YYYY
                    date_match = re.search(r'(\d{2}/\d{2}/\d{2,4})', cps_text)
                    if date_match:
                        date_str = date_match.group(1)
                        fmt = '%d/%m/%Y' if len(date_str.split('/')[-1]) == 4 else '%d/%m/%y'
                        fecha_venc = datetime.strptime(date_str, fmt).date()
                        tipo_cps_default, _ = TipoCps.objects.get_or_create(codigo='DT9', defaults={'descripcion': 'Certificación DT9'})
                        
                        OrganizacionCps.objects.update_or_create(
                            organizacion=org,
                            codigo='REGISTRO-EXCEL',
                            defaults={
                                'fecha_vencimiento': fecha_venc,
                                'fecha_expedicion': datetime.now().date(),
                                'tipo_cps': tipo_cps_default,
                                'activa': True
                            }
                        )

            return Response({
                "message": f"Importación completada: {created_count} creados, {updated_count} actualizados.",
                "created": created_count,
                "updated": updated_count
            })

        except Exception as e:
            return Response({"error": f"Error al procesar el archivo: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
