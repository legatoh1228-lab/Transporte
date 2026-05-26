from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from django.db import transaction
from django.utils import timezone
from catalogs.models import Modalidad, SubModalidad, TipoCombustible, TerritorioMunicipio, TipoOrganizacion, TipoCps
from organizations.models import EmpresaOrganizacion
from .models import FlotaVehiculo, VehiculoOrganizacion, Terminal, VehiculoOperador, AsignacionRuta
from .serializers import (
    FlotaVehiculoSerializer, VehiculoOrganizacionSerializer, 
    TerminalSerializer, VehiculoOperadorSerializer, AsignacionRutaSerializer
)

class AsignacionRutaViewSet(viewsets.ModelViewSet):
    queryset = AsignacionRuta.objects.all().order_by('-fecha_inicio')
    serializer_class = AsignacionRutaSerializer
    filterset_fields = ['vehiculo', 'operador', 'horario', 'estatus']

    def perform_create(self, serializer):
        from django.core.exceptions import ValidationError as DjangoValidationError
        from rest_framework.exceptions import ValidationError as DRFValidationError
        try:
            serializer.save()
        except DjangoValidationError as e:
            raise DRFValidationError(e.message_dict if hasattr(e, 'message_dict') else e.messages)

    def perform_update(self, serializer):
        from django.core.exceptions import ValidationError as DjangoValidationError
        from rest_framework.exceptions import ValidationError as DRFValidationError
        try:
            serializer.save()
        except DjangoValidationError as e:
            raise DRFValidationError(e.message_dict if hasattr(e, 'message_dict') else e.messages)

class TerminalViewSet(viewsets.ModelViewSet):
    queryset = Terminal.objects.all()
    serializer_class = TerminalSerializer
    filterset_fields = ['municipio', 'tipo', 'estatus']
    search_fields = ['nombre']

class FlotaVehiculoViewSet(viewsets.ModelViewSet):
    queryset = FlotaVehiculo.objects.all()
    serializer_class = FlotaVehiculoSerializer
    filterset_fields = ['modalidad', 'submodalidad', 'marca']
    search_fields = ['placa', 'marca', 'modelo']

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    @transaction.atomic
    def bulk_delete(self, request):
        placas = request.data.get('placas', [])
        delete_all = request.data.get('all', False)
        
        if delete_all:
            count, _ = FlotaVehiculo.objects.all().delete()
            return Response({'message': f'Se han eliminado todas las {count} unidades de la flota correctamente.'}, status=status.HTTP_200_OK)
            
        if not placas:
            return Response({'error': 'No se proporcionaron placas para eliminar.'}, status=status.HTTP_400_BAD_REQUEST)
            
        deleted_count, _ = FlotaVehiculo.objects.filter(placa__in=placas).delete()
        return Response({'message': f'Se han eliminado {deleted_count} unidades correctamente.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='import-excel')
    @transaction.atomic
    def import_excel(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No se proporcionó ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            # Find "REGISTRO FLOTA" sheet
            sheet_name = 'REGISTRO FLOTA'
            if sheet_name not in wb.sheetnames:
                # Look for matching sheet name
                for name in wb.sheetnames:
                    if 'flota' in name.lower():
                        sheet_name = name
                        break
                else:
                    return Response({'error': 'No se encontró la hoja "REGISTRO FLOTA" en el archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)
            
            sheet = wb[sheet_name]
            
            # Resolve static modals/combustibles/etc
            modalidad_colectivo = Modalidad.objects.filter(nombre='COLECTIVO').first()
            if not modalidad_colectivo:
                modalidad_colectivo = Modalidad.objects.first() # Fallback

            submodalidad_minibus = SubModalidad.objects.filter(nombre__icontains='minibús').first() or SubModalidad.objects.filter(nombre__icontains='camionetica').first()
            submodalidad_bus = SubModalidad.objects.filter(nombre__icontains='autobús').first() or SubModalidad.objects.filter(nombre__icontains='bus grande').first()
            submodalidad_buseta = SubModalidad.objects.filter(nombre__icontains='buseta').first()
            
            # If not found, use defaults or fallback
            if not submodalidad_minibus: submodalidad_minibus = SubModalidad.objects.first()
            if not submodalidad_bus: submodalidad_bus = SubModalidad.objects.first()
            if not submodalidad_buseta: submodalidad_buseta = SubModalidad.objects.first()
            
            combustibles = {
                'gasolina': TipoCombustible.objects.filter(nombre='GASOLINA').first() or TipoCombustible.objects.first(),
                'diesel': TipoCombustible.objects.filter(nombre__in=['GASOIL', 'DIESEL']).first() or TipoCombustible.objects.first(),
                'gas': TipoCombustible.objects.filter(nombre='GAS').first() or TipoCombustible.objects.first(),
            }

            created_count = 0
            updated_count = 0
            errors = []
            
            # Loop starting from row 9 (1-indexed)
            # Headers are on row 8
            for row_idx, row in enumerate(sheet.iter_rows(min_row=9, values_only=True), start=9):
                # Skip empty rows or rows that do not have RIF and Plates
                if not any(row):
                    continue
                
                municipio_name = row[0]
                org_name = row[1]
                rif = row[2]
                
                minibus_plate = row[3]
                bus_plate = row[4]
                otro_plate = row[5]
                
                cps_code = row[6]
                
                gasolina_mark = row[7]
                diesel_mark = row[8]
                gas_mark = row[9]
                
                observacion = row[10]

                # Determine the plate and submodalidad
                placa = None
                submod = None
                
                if minibus_plate:
                    placa = str(minibus_plate).strip()
                    submod = submodalidad_minibus
                elif bus_plate:
                    placa = str(bus_plate).strip()
                    submod = submodalidad_bus
                elif otro_plate:
                    placa = str(otro_plate).strip()
                    submod = submodalidad_buseta
                
                if not placa:
                    # Skip rows that don't have a plate
                    continue
                
                if not rif:
                    errors.append(f"Fila {row_idx}: Falta el RIF de la organización.")
                    continue

                # Ensure organization exists or create it
                rif = str(rif).strip()
                org_name = str(org_name or f"Línea RIF {rif}").strip()
                
                try:
                    # Wrap each row's database operations in a transaction savepoint (atomic block)
                    with transaction.atomic():
                        municipio = None
                        if municipio_name:
                            municipio_name = str(municipio_name).strip().upper()
                            # Look up TerritorioMunicipio
                            municipio = TerritorioMunicipio.objects.filter(nombre__iexact=municipio_name).first()
                        
                        if not municipio:
                            # Default to Sucre if it's "SUCRE", otherwise first municipio
                            municipio = TerritorioMunicipio.objects.filter(nombre__iexact='SUCRE').first() or TerritorioMunicipio.objects.first()

                        tipo_org = TipoOrganizacion.objects.filter(nombre__icontains='LÍNEA').first() or TipoOrganizacion.objects.first()
                        org, org_created = EmpresaOrganizacion.objects.get_or_create(
                            rif=rif,
                            defaults={
                                'razon_social': org_name,
                                'municipio': municipio,
                                'tipo': tipo_org,
                                'fecha_constitucion_mercantil': timezone.now().date(),
                                'cupo_maximo_unidades': 100,
                            }
                        )

                        # Si el vehículo tiene un CPS específico (ej. DT9/1 o DT10/2), garantizamos que la organización tenga ese tipo de CPS registrado
                        if cps_code:
                            cps_code_clean = str(cps_code).strip().upper()
                            prefix = None
                            if 'DT9' in cps_code_clean:
                                prefix = 'DT9'
                            elif 'DT10' in cps_code_clean:
                                prefix = 'DT10'
                            
                            if prefix:
                                tipo_cps, _ = TipoCps.objects.get_or_create(
                                    codigo=prefix,
                                    defaults={'descripcion': f'Certificación {prefix}'}
                                )
                                from organizations.models import OrganizacionCps
                                OrganizacionCps.objects.get_or_create(
                                    organizacion=org,
                                    tipo_cps=tipo_cps,
                                    defaults={
                                        'codigo': f'CPS-{prefix}-{rif}',
                                        'fecha_expedicion': timezone.now().date(),
                                        'fecha_vencimiento': timezone.now().date() + timezone.timedelta(days=365),
                                        'activa': True
                                    }
                                )

                        # Resolve combustible
                        comb = combustibles['gasolina']
                        if diesel_mark and str(diesel_mark).strip().upper() == 'X':
                            comb = combustibles['diesel']
                        elif gas_mark and str(gas_mark).strip().upper() == 'X':
                            comb = combustibles['gas']
                        elif gasolina_mark and str(gasolina_mark).strip().upper() == 'X':
                            comb = combustibles['gasolina']

                        # Create or update FlotaVehiculo
                        vehicle_defaults = {
                            'modalidad': modalidad_colectivo,
                            'submodalidad': submod,
                            'combustible': comb,
                            'cps': cps_code,
                            'observaciones': observacion,
                            'marca': 'No especificado',
                            'modelo': 'No especificado',
                            'anio': 2026,
                            'capacidad': 32 if submod == submodalidad_bus else (15 if submod == submodalidad_minibus else 5),
                        }

                        vehicle, v_created = FlotaVehiculo.objects.update_or_create(
                            placa=placa,
                            defaults=vehicle_defaults
                        )

                        # Link to Organization if not linked
                        VehiculoOrganizacion.objects.get_or_create(
                            vehiculo=vehicle,
                            organizacion=org,
                            defaults={
                                'fecha_inicio': timezone.now().date()
                            }
                        )

                        if v_created:
                            created_count += 1
                        else:
                            updated_count += 1

                except Exception as row_err:
                    errors.append(f"Fila {row_idx} (Placa {placa}): {str(row_err)}")

            return Response({
                'message': 'Importación completada exitosamente.',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': f'Error al procesar el archivo Excel: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class VehiculoOrganizacionViewSet(viewsets.ModelViewSet):
    queryset = VehiculoOrganizacion.objects.all()
    serializer_class = VehiculoOrganizacionSerializer
    filterset_fields = ['organizacion', 'vehiculo']

class VehiculoOperadorViewSet(viewsets.ModelViewSet):
    queryset = VehiculoOperador.objects.all()
    serializer_class = VehiculoOperadorSerializer
    
    def perform_create(self, serializer):
        # Desactivar asignaciones previas para este vehículo
        vehiculo = serializer.validated_data.get('vehiculo')
        VehiculoOperador.objects.filter(vehiculo=vehiculo, estatus='Activo').update(estatus='Inactivo')
        serializer.save(estatus='Activo')

